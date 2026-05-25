"""Generate src/config/qc1Catalog.ts from AC Excel template. Run from repo root."""
import json
import re
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
AC_XLSX = Path(r"c:\Users\HP\Desktop\Template Name - Area - AC Quote 1 - New.xlsx")
OUT = REPO / "src" / "config" / "qc1Catalog.ts"


def slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")[:50]


def infer_tags(cat: str, name: str, part: str) -> list[str]:
    t: list[str] = []
    blob = f"{cat} {name} {part}".upper()
    if "R410" in blob:
        t.append("refrigerant:r410")
    if "R32" in blob:
        t.append("refrigerant:r32")
    if "CASSETTE" in blob or "CASSETE" in blob:
        t.append("unit-type:cassette")
    if "MID WALL" in blob or "MIDWALL" in blob.replace(" ", ""):
        t.append("unit-type:mid-wall")
    if "DUCTED" in blob or "DUCT" in blob or "CONCEALED" in blob or "HIDE AWAY" in blob:
        t.append("unit-type:ducted")
    if "/3" in part or "3 PHASE" in blob:
        t.append("phase:three")
    elif any(x in blob for x in ("SPLIT", "CASSETTE", "INVERTER", "BTU")):
        t.append("phase:single")
    if "ALLIANCE" in blob or "COMFEE" in blob or part.startswith(("FOU", "FCM")):
        t.append("brand:alliance")
    if "SAMSUNG" in blob:
        t.append("brand:samsung")
    if "HISENSE" in blob:
        t.append("brand:hisense")
    if "PIPING" in blob or "TUBING" in blob or "TUBE" in blob:
        t.append("line:refrigerant-piping")
    if "GRILL" in blob or "DIFFUSER" in blob or "DUCTING" in blob:
        t.append("line:air-distribution")
    return t


def extract_ac():
    wb = openpyxl.load_workbook(AC_XLSX, data_only=True)
    ws = wb["QC1"]
    current_cat = None
    current_id = None
    cats = []
    products = []
    for r in range(22, 280):
        a, b, c, e = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 5).value
        if b and isinstance(b, str) and (not a or str(a).strip() == "") and len(b.strip()) > 12:
            blob = b.strip().replace("\n", " ")
            if blob.isupper() or any(
                x in blob for x in ("ALLIANCE", "SAMSUNG", "HISENSE", "DUCT", "CASSETTE", "GRILL", "FLEXIBLE", "DIFFUSER")
            ):
                current_cat = blob
                current_id = f"ac-{slug(blob)}"
                if not any(x["id"] == current_id for x in cats):
                    cats.append({"id": current_id, "label": blob, "templateId": "ac_v1"})
        elif current_id and b and e and a:
            try:
                cost = float(e)
            except (TypeError, ValueError):
                continue
            part = str(a).strip()
            name = str(b).strip()
            products.append(
                {
                    "id": f"ac-{slug(part)}-{r}",
                    "templateId": "ac_v1",
                    "categoryId": current_id,
                    "partCode": part,
                    "name": name[:100],
                    "comment": str(c).strip() if c else None,
                    "listPrice": round(cost, 2),
                    "tags": infer_tags(current_cat or "", name, part),
                }
            )
    wb.close()
    return cats, products


def main():
    ac_cats, ac_products = extract_ac()
    # Solar seed — mapped to QC1 column A categories (PV template)
    from pathlib import Path as P

    solar_path = REPO / "src" / "config" / "solarQuoteProducts.ts"
    # Solar products stay in solarQuoteProducts.ts; catalog merges at runtime in TS we generate

    header = '''/** Auto-generated AC catalog from Template Name - Area - AC Quote 1 - New.xlsx. Regenerate: python scripts/utilities/gen_qc1_catalog.py */
import type { Qc1Category, QuoteProduct, QuoteTemplateId } from "@/types/quotes";
import { SOLAR_QC1_CATEGORIES, SOLAR_QC1_PRODUCTS } from "@/config/solarQc1Products";

export const QUOTE_TEMPLATES: { id: QuoteTemplateId; label: string; description: string }[] = [
  { id: "ac_v1", label: "Air conditioning", description: "QC1 → Quote (Alliance / Samsung / Hisense)" },
  { id: "solar_pv_v1", label: "Solar PV", description: "QC1 → Quote + savings calculator" },
];

export const AC_QC1_CATEGORIES: Qc1Category[] = '''
    footer = '''

export const AC_QC1_PRODUCTS: QuoteProduct[] = '''

    templates_footer = '''

export function getCategories(templateId: QuoteTemplateId): Qc1Category[] {
  if (templateId === "ac_v1") return AC_QC1_CATEGORIES;
  return SOLAR_QC1_CATEGORIES;
}

export function getProducts(templateId: QuoteTemplateId): QuoteProduct[] {
  if (templateId === "ac_v1") return AC_QC1_PRODUCTS;
  return SOLAR_QC1_PRODUCTS;
}

export const ALL_QC1_PRODUCTS: QuoteProduct[] = [...AC_QC1_PRODUCTS, ...SOLAR_QC1_PRODUCTS];
'''

    OUT.write_text(
        header + json.dumps(ac_cats, indent=2) + " as const satisfies Qc1Category[];\n\n"
        + footer + json.dumps(ac_products, indent=2) + " as unknown as QuoteProduct[];\n"
        + templates_footer,
        encoding="utf-8",
    )
    print(f"Wrote {OUT} ({len(ac_cats)} categories, {len(ac_products)} products)")


if __name__ == "__main__":
    main()
